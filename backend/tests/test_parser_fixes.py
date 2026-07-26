import io
import pytest
import openpyxl
from app.services.excel_engine.erp_excel_parser import ERPExcelParser
from app.services.doc_parser import document_parser

def create_stream(wb: openpyxl.Workbook) -> bytes:
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()

def test_employees_on_leave_and_operational_issues_parsing():
    """Test extracting 'Employees on Leave' and 'Operational Issues' placed after employee performance table."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pothys Daily Report"

    # Branch summary top
    ws.cell(row=2, column=1, value="Branch Name")
    ws.cell(row=2, column=2, value="Chromepet Swarna Mahal")
    ws.cell(row=3, column=1, value="Total Revenue")
    ws.cell(row=3, column=2, value=450000.0)

    # Employee table
    headers = [
        "Employee Name", "Department", "Gold Grams", "Gold Sales",
        "Silver Grams", "Silver Sales", "DigiGold", "DigiSilver"
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=5, column=c, value=h)

    # Add employees
    ws.cell(row=6, column=1, value="Karthik Raj")
    ws.cell(row=6, column=2, value="Sales Executive")
    ws.cell(row=6, column=3, value=15.0)
    ws.cell(row=6, column=4, value=110000.0)
    ws.cell(row=6, column=5, value=200.0)
    ws.cell(row=6, column=6, value=25000.0)
    ws.cell(row=6, column=7, value=6)
    ws.cell(row=6, column=8, value=9)

    # Directly below table (row 8, no 3-row gap): HR leave and Operational Issues
    ws.cell(row=8, column=1, value="Employees Present")
    ws.cell(row=8, column=2, value=15)

    ws.cell(row=9, column=1, value="Employees on Leave")
    ws.cell(row=9, column=2, value=3)

    ws.cell(row=10, column=1, value="Operational Issues")
    ws.cell(row=10, column=2, value="UPS backup fault on 2nd floor counter.")

    res = ERPExcelParser.parse(create_stream(wb))
    summary = res["summary"]

    assert summary["branch_name"] == "Chromepet Swarna Mahal"
    assert summary["employees_present"] == 15
    assert summary["employees_absent"] == 3
    assert summary["operational_issues"] == "UPS backup fault on 2nd floor counter."

    # Verify complete employee record structure
    employees = res["employees"]
    assert len(employees) == 1
    emp = employees[0]
    assert emp["employee_name"] == "Karthik Raj"
    assert emp["department"] == "Sales Executive"
    assert emp["sales"] == 135000.0
    assert emp["digigold"] == 6
    assert emp["digisilver"] == 9
    assert emp["gold_grams"] == 15.0
    assert emp["silver_grams"] == 200.0

def test_inline_key_value_extraction():
    """Test extracting inline key-value pairs like 'Employees on Leave: 2'."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    ws.cell(row=2, column=1, value="Branch Name: Poonamallee Swarna Mahal")
    ws.cell(row=3, column=1, value="Total Revenue: 500000")
    ws.cell(row=4, column=1, value="Employees on Leave: 2 staff")
    ws.cell(row=5, column=1, value="Operational Issues - HVAC cooling glitch in billing area")

    res = ERPExcelParser.parse(create_stream(wb))
    summary = res["summary"]

    assert summary["branch_name"] == "Poonamallee Swarna Mahal"
    assert summary["total_revenue"] == 500000.0
    assert summary["employees_absent"] == 2
    assert summary["operational_issues"] == "HVAC cooling glitch in billing area"

def test_employees_on_leave_equals_syntax():
    """Test extracting 'Employees Present = 57' and 'Employees on Leave = 2'."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pothys Daily Report"

    ws.cell(row=2, column=1, value="Branch Name")
    ws.cell(row=2, column=2, value="T.Nagar Swarna Mahal")
    ws.cell(row=3, column=1, value="Total Revenue")
    ws.cell(row=3, column=2, value="570000")
    ws.cell(row=4, column=1, value="Employees Present = 57")
    ws.cell(row=5, column=1, value="Employees on Leave = 2")

    res = ERPExcelParser.parse(create_stream(wb))
    summary = res["summary"]

    assert summary["employees_present"] == 57
    assert summary["employees_absent"] == 2
