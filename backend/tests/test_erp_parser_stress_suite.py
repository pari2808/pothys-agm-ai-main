import io
import pytest
import openpyxl
from openpyxl.styles import Font, PatternFill, Side, Border
from app.services.excel_engine.erp_excel_parser import ERPExcelParser

def create_stream(wb: openpyxl.Workbook) -> bytes:
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()

# ── 1. Extra blank rows ──
def test_01_extra_blank_rows():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=12, column=2, value="Branch Name")
    ws.cell(row=12, column=3, value="Chennai Swarna Mahal")
    ws.cell(row=18, column=2, value="Gold Sales")
    ws.cell(row=18, column=3, value=120000.0)
    ws.cell(row=25, column=2, value="Total Revenue")
    ws.cell(row=25, column=3, value=120000.0)

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["branch_name"] == "Chennai Swarna Mahal"
    assert res["summary"]["gold_sales"] == 120000.0
    assert res["diagnostics"]["confidence_score"] >= 0.10

# ── 2. Extra blank columns ──
def test_02_extra_blank_columns():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=3, column=6, value="Location")
    ws.cell(row=3, column=7, value="Velachery Swarna Mahal")
    ws.cell(row=5, column=6, value="Gold Sales")
    ws.cell(row=5, column=7, value=90000.0)

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["branch_name"] == "Velachery Swarna Mahal"
    assert res["summary"]["gold_sales"] == 90000.0

# ── 3. Columns in random order ──
def test_03_random_column_order():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"
    headers = ["DigiSilver Count", "Gold Sales", "Staff Name", "Role", "Gold Weight (g)"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    ws.cell(row=2, column=1, value=5)
    ws.cell(row=2, column=2, value=45000.0)
    ws.cell(row=2, column=3, value="Karthik")
    ws.cell(row=2, column=4, value="Executive")
    ws.cell(row=2, column=5, value=6.2)

    res = ERPExcelParser.parse(create_stream(wb))
    emps = res["employees"]
    assert len(emps) == 1
    assert emps[0]["name"] == "Karthik"
    assert emps[0]["gold_amount"] == 45000.0
    assert emps[0]["digisilver_enrollments"] == 5

# ── 4. Tables starting at random positions ──
def test_04_random_table_positions():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=15, column=8, value="Branch Name")
    ws.cell(row=15, column=9, value="Salem Swarna Mahal")
    ws.cell(row=17, column=8, value="Total Sales")
    ws.cell(row=17, column=9, value=300000.0)

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["branch_name"] == "Salem Swarna Mahal"
    assert res["summary"]["total_revenue"] == 300000.0

# ── 5. Different worksheet names ──
def test_05_different_worksheet_names():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Operations Log 2026"
    ws.cell(row=2, column=1, value="Showroom")
    ws.cell(row=2, column=2, value="Tirunelveli Mahal")
    ws.cell(row=4, column=1, value="Gold Sales")
    ws.cell(row=4, column=2, value=75000.0)

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["branch_name"] == "Tirunelveli Mahal"
    assert res["summary"]["gold_sales"] == 75000.0

# ── 6. Multiple summary sheets ──
def test_06_multiple_summary_sheets():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Summary Part 1"
    ws1.cell(row=2, column=1, value="Branch")
    ws1.cell(row=2, column=2, value="Pondicherry Mahal")

    ws2 = wb.create_sheet(title="Summary Part 2")
    ws2.cell(row=2, column=1, value="Gold Sales")
    ws2.cell(row=2, column=2, value=110000.0)

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["branch_name"] == "Pondicherry Mahal"
    assert res["summary"]["gold_sales"] == 110000.0

# ── 7. Different header aliases ──
def test_07_different_header_aliases():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=2, column=1, value="Today's Gold Collection")
    ws.cell(row=2, column=2, value=88000.0)
    ws.cell(row=3, column=1, value="Staff Headcount Present")
    ws.cell(row=3, column=2, value=22)

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["gold_sales"] == 88000.0
    assert res["summary"]["employees_present"] == 22

# ── 8. Merged cells ──
def test_08_merged_cells():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.merge_cells("A5:C5")
    ws.cell(row=5, column=1, value="Gold Sales")
    ws.merge_cells("D5:F5")
    ws.cell(row=5, column=4, value=135000.0)

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["gold_sales"] == 135000.0

# ── 9. Unmerged cells ──
def test_09_unmerged_cells():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Branch")
    ws.cell(row=1, column=2, value="Erode Swarna Mahal")
    ws.cell(row=2, column=1, value="Total Revenue")
    ws.cell(row=2, column=2, value=50000.0)

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["branch_name"] == "Erode Swarna Mahal"
    assert res["summary"]["total_revenue"] == 50000.0

# ── 10. Hidden rows ──
def test_10_hidden_rows():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=5, column=1, value="Gold Sales")
    ws.cell(row=5, column=2, value=65000.0)
    ws.row_dimensions[5].hidden = True

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["gold_sales"] == 65000.0

# ── 11. Hidden columns ──
def test_11_hidden_columns():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=2, column=2, value="Branch Name")
    ws.cell(row=2, column=3, value="Thanjavur Mahal")
    ws.column_dimensions["B"].hidden = True

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["branch_name"] == "Thanjavur Mahal"

# ── 12. Different fonts ──
def test_12_different_fonts():
    wb = openpyxl.Workbook()
    ws = wb.active
    c1 = ws.cell(row=2, column=1, value="Gold Sales")
    c1.font = Font(name="Comic Sans MS", size=18, bold=True)
    c2 = ws.cell(row=2, column=2, value=92000.0)
    c2.font = Font(name="Impact", size=10)

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["gold_sales"] == 92000.0

# ── 13. Different colors ──
def test_13_different_colors():
    wb = openpyxl.Workbook()
    ws = wb.active
    fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    c = ws.cell(row=2, column=1, value="Gold Sales")
    c.fill = fill
    ws.cell(row=2, column=2, value=105000.0)

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["gold_sales"] == 105000.0

# ── 14. Borders removed ──
def test_14_borders_removed():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.views.sheetView[0].showGridLines = False
    ws.cell(row=2, column=1, value="Silver Sales")
    ws.cell(row=2, column=2, value=34000.0)

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["silver_sales"] == 34000.0

# ── 15. Additional irrelevant columns ──
def test_15_additional_irrelevant_columns():
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["Emp Name", "Store Code", "Designation", "KPI Rating", "Gold Sales", "Commission %"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    ws.cell(row=2, column=1, value="Dhanush")
    ws.cell(row=2, column=2, value="MAS-01")
    ws.cell(row=2, column=3, value="Senior Executive")
    ws.cell(row=2, column=4, value="A+")
    ws.cell(row=2, column=5, value=140000.0)
    ws.cell(row=2, column=6, value="2.5%")

    res = ERPExcelParser.parse(create_stream(wb))
    emps = res["employees"]
    assert len(emps) == 1
    assert emps[0]["name"] == "Dhanush"
    assert emps[0]["gold_amount"] == 140000.0

# ── 16. Missing optional fields ──
def test_16_missing_optional_fields():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Gold Sales")
    ws.cell(row=1, column=2, value=50000.0)

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["gold_sales"] == 50000.0
    assert res["summary"]["platinum_sales"] == 0.0
    assert res["summary"]["remarks"] == "None"
    assert res["diagnostics"]["status"] in ["SUCCESS", "PARTIAL_SUCCESS", "CRITICAL_MISSING"]

# ── 17. Missing required fields ──
def test_17_missing_required_fields():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Unrelated Label")
    ws.cell(row=1, column=2, value="Unrelated Value")

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["diagnostics"]["confidence_score"] < 0.5
    assert len(res["diagnostics"]["warnings"]) > 0

# ── 18. Currency symbols ──
def test_18_currency_symbols():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=2, column=1, value="Gold Sales")
    ws.cell(row=2, column=2, value="₹ 1,50,000.00")
    ws.cell(row=3, column=1, value="Silver Sales")
    ws.cell(row=3, column=2, value="Rs. 25,000/-")
    ws.cell(row=4, column=1, value="Total Revenue")
    ws.cell(row=4, column=2, value="INR 175000")

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["gold_sales"] == 150000.0
    assert res["summary"]["silver_sales"] == 25000.0
    assert res["summary"]["total_revenue"] == 175000.0

# ── 19. Numbers stored as text ──
def test_19_numbers_stored_as_text():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=2, column=1, value="Gold Sales")
    ws.cell(row=2, column=2, value="95000.00")
    ws.cell(row=3, column=1, value="Employees Present")
    ws.cell(row=3, column=2, value="15")

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["gold_sales"] == 95000.0
    assert res["summary"]["employees_present"] == 15

# ── 20. Formulas instead of values ──
def test_20_formulas_instead_of_values():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Gold Sales")
    ws.cell(row=1, column=2, value=100000.0)
    ws.cell(row=2, column=1, value="Silver Sales")
    ws.cell(row=2, column=2, value=50000.0)
    ws.cell(row=3, column=1, value="Total Revenue")
    ws.cell(row=3, column=2, value="=SUM(B1:B2)")

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["gold_sales"] == 100000.0
    assert res["summary"]["silver_sales"] == 50000.0
    assert res["summary"]["total_revenue"] == 150000.0

# ── 21. Multiple employee tables ──
def test_21_multiple_employee_tables():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Shift 1"
    ws1.cell(row=1, column=1, value="Emp Name")
    ws1.cell(row=1, column=2, value="Gold Sales")
    ws1.cell(row=2, column=1, value="Kamal")
    ws1.cell(row=2, column=2, value=60000.0)

    ws2 = wb.create_sheet(title="Shift 2")
    ws2.cell(row=1, column=1, value="Emp Name")
    ws2.cell(row=1, column=2, value="Gold Sales")
    ws2.cell(row=2, column=1, value="Rajini")
    ws2.cell(row=2, column=2, value=80000.0)

    res = ERPExcelParser.parse(create_stream(wb))
    emps = res["employees"]
    assert len(emps) == 2
    names = [e["name"] for e in emps]
    assert "Kamal" in names
    assert "Rajini" in names

# ── 22. Extra notes between tables ──
def test_22_extra_notes_between_tables():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Gold Sales")
    ws.cell(row=1, column=2, value=40000.0)
    ws.cell(row=4, column=1, value="* NOTE: All sales figures are verified by audit department.")
    ws.cell(row=8, column=1, value="Staff Name")
    ws.cell(row=8, column=2, value="Gold Sales")
    ws.cell(row=9, column=1, value="Suriya")
    ws.cell(row=9, column=2, value=40000.0)

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["gold_sales"] == 40000.0
    assert len(res["employees"]) == 1
    assert res["employees"][0]["name"] == "Suriya"

# ── 23. Attendance table before summary ──
def test_23_attendance_table_before_summary():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=2, column=1, value="Staff Name")
    ws.cell(row=2, column=2, value="Gold Sales")
    ws.cell(row=3, column=1, value="Ajith")
    ws.cell(row=3, column=2, value=99000.0)

    ws.cell(row=15, column=1, value="Location")
    ws.cell(row=15, column=2, value="Vellore Swarna Mahal")
    ws.cell(row=16, column=1, value="Total Sales")
    ws.cell(row=16, column=2, value=99000.0)

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["branch_name"] == "Vellore Swarna Mahal"
    assert len(res["employees"]) == 1
    assert res["employees"][0]["name"] == "Ajith"

# ── 24. Summary after attendance ──
def test_24_summary_after_attendance():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Team Attendance"
    ws1.cell(row=1, column=1, value="Staff Name")
    ws1.cell(row=1, column=2, value="Gold Sales")
    ws1.cell(row=2, column=1, value="Vikram")
    ws1.cell(row=2, column=2, value=115000.0)

    ws2 = wb.create_sheet(title="Executive Summary")
    ws2.cell(row=1, column=1, value="Branch Name")
    ws2.cell(row=1, column=2, value="Kanchipuram Mahal")

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["branch_name"] == "Kanchipuram Mahal"
    assert len(res["employees"]) == 1
    assert res["employees"][0]["name"] == "Vikram"

# ── 25. Mixed capitalization ──
def test_25_mixed_capitalization():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=2, column=1, value="gOlD sAlEs")
    ws.cell(row=2, column=2, value=123000.0)
    ws.cell(row=3, column=1, value="StAfF pReSeNt")
    ws.cell(row=3, column=2, value=14)

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["gold_sales"] == 123000.0
    assert res["summary"]["employees_present"] == 14

# ── 26. Leading/trailing spaces ──
def test_26_leading_trailing_spaces():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=2, column=1, value="   Gold Sales \t\n")
    ws.cell(row=2, column=2, value=77000.0)
    ws.cell(row=3, column=1, value="  Branch Name  ")
    ws.cell(row=3, column=2, value="  Hosur Swarna Mahal  ")

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["gold_sales"] == 77000.0
    assert res["summary"]["branch_name"] == "Hosur Swarna Mahal"

# ── 27. Unicode characters ──
def test_27_unicode_characters():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=2, column=1, value="போத்தீஸ் Gold Sales")
    ws.cell(row=2, column=2, value=160000.0)
    ws.cell(row=3, column=1, value="Manager Remarks")
    ws.cell(row=3, column=2, value="All good 👍 - Всё отлично")

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["gold_sales"] == 160000.0
    assert "All good" in res["summary"]["remarks"]

# ── 28. Empty worksheets ──
def test_28_empty_worksheets():
    wb = openpyxl.Workbook()
    wb.create_sheet("Blank1")
    wb.create_sheet("Blank2")
    ws = wb.active
    ws.title = "Data"
    ws.cell(row=2, column=1, value="Gold Sales")
    ws.cell(row=2, column=2, value=55000.0)

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["gold_sales"] == 55000.0

# ── 29. Multiple unrelated worksheets ──
def test_29_multiple_unrelated_worksheets():
    wb = openpyxl.Workbook()
    ws_unrelated = wb.active
    ws_unrelated.title = "Vendor Contacts"
    ws_unrelated.cell(row=1, column=1, value="Supplier Name")
    ws_unrelated.cell(row=1, column=2, value="Phone Number")

    ws_data = wb.create_sheet(title="Report Summary")
    ws_data.cell(row=2, column=1, value="Total Revenue")
    ws_data.cell(row=2, column=2, value=210000.0)

    res = ERPExcelParser.parse(create_stream(wb))
    assert res["summary"]["total_revenue"] == 210000.0

# ── 30. Duplicate headers ──
def test_30_duplicate_headers():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Staff Name")
    ws.cell(row=1, column=2, value="Gold Sales")
    ws.cell(row=2, column=1, value="Nayan")
    ws.cell(row=2, column=2, value=50000.0)

    ws.cell(row=10, column=1, value="Staff Name")
    ws.cell(row=10, column=2, value="Gold Sales")
    ws.cell(row=11, column=1, value="Trisha")
    ws.cell(row=11, column=2, value=70000.0)

    res = ERPExcelParser.parse(create_stream(wb))
    emps = res["employees"]
    assert len(emps) == 2
    names = [e["name"] for e in emps]
    assert "Nayan" in names
    assert "Trisha" in names
