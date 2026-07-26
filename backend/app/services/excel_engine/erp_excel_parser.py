"""
erp_excel_parser.py — Production Excel Parser Orchestrator
==========================================================

Coordinates metadata extraction + employee table parsing for the new
Pothys manager template format.

Output structure:
  {
    "summary": { report_date, branch_name, manager_name, ..., gold, diamond, ... },
    "employees": [ { name, gold, diamond, platinum, silver, silver_mrp, ... }, ... ],
    "scheme_summary": { subhiksham_count, subhiksham_value, viruksham_count, viruksham_value, ... },
    "diagnostics": { ... }
  }
"""

import io
import openpyxl
from typing import Dict, Any
from app.services.excel_engine.normalizer import SheetMatrix
from app.services.excel_engine.sheet_classifier import SheetClassifier
from app.services.excel_engine.anchor_parser import AnchorParser
from app.services.excel_engine.table_parser import TableParser
from app.services.excel_engine.validator import ExtractionValidator


class ERPExcelParser:
    """Production Excel parser for Pothys daily branch reports — new template format."""

    @staticmethod
    def parse(file_content: bytes) -> Dict[str, Any]:
        """
        Parse Excel workbook bytes from the new production manager template.

        The new template has a flat employee table with columns:
          Employee Name | Gold | Diamond | Platinum | Silver | Silver MRP |
          Subhiksham Count | Subhiksham Value | Viruksham Count | Viruksham Value |
          DigiGold | DigiSilver

        Branch-level summary values (Gold, Diamond, etc.) are AGGREGATED from
        employee rows — not parsed from separate scalar key-value cells.
        """
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        except Exception as e:
            raise ValueError(f"Invalid Excel file format: {str(e)}")

        if not wb.sheetnames:
            raise ValueError("Workbook contains no sheets.")

        # 1. Wrap all sheets in SheetMatrix to handle merged cells transparently
        matrices = [SheetMatrix(wb[name]) for name in wb.sheetnames]

        # 2. Classify sheets
        classified = SheetClassifier.classify_workbook(matrices)

        # 3. Extract metadata (date, branch, manager, attendance, complaints, issues, remarks)
        #    Search all sheets for metadata labels
        summary_data = AnchorParser.extract_summary(matrices)

        # 4. Extract employee records from data sheets (or all sheets as fallback)
        data_matrices = classified["data"] if classified["data"] else matrices
        employees_data = TableParser.extract_employee_records(data_matrices)

        # If no employees found in data sheets, try all sheets
        if not employees_data and classified["data"]:
            employees_data = TableParser.extract_employee_records(matrices)

        # 5. Aggregate branch-level summary from employee rows
        total_gold = sum(e.get("gold", 0) for e in employees_data)
        total_diamond = sum(e.get("diamond", 0) for e in employees_data)
        total_platinum = sum(e.get("platinum", 0) for e in employees_data)
        total_silver = sum(e.get("silver", 0) for e in employees_data)
        total_silver_mrp = sum(e.get("silver_mrp", 0) for e in employees_data)
        total_subhiksham_count = sum(e.get("subhiksham_count", 0) for e in employees_data)
        total_subhiksham_value = sum(e.get("subhiksham_value", 0) for e in employees_data)
        total_viruksham_count = sum(e.get("viruksham_count", 0) for e in employees_data)
        total_viruksham_value = sum(e.get("viruksham_value", 0) for e in employees_data)
        total_digigold = sum(e.get("digigold", 0) for e in employees_data)
        total_digisilver = sum(e.get("digisilver", 0) for e in employees_data)
        total_revenue = float(total_gold + total_diamond + total_platinum + total_silver)

        # 6. Build scheme summary from aggregated Subhiksham / Viruksham data
        scheme_items = []
        if total_subhiksham_count > 0 or total_subhiksham_value > 0:
            scheme_items.append({
                "scheme_name": "Subhiksham",
                "count": int(total_subhiksham_count),
                "value": float(total_subhiksham_value),
            })
        if total_viruksham_count > 0 or total_viruksham_value > 0:
            scheme_items.append({
                "scheme_name": "Viruksham",
                "count": int(total_viruksham_count),
                "value": float(total_viruksham_value),
            })

        scheme_summary = {
            "subhiksham_count": int(total_subhiksham_count),
            "subhiksham_value": float(total_subhiksham_value),
            "viruksham_count": int(total_viruksham_count),
            "viruksham_value": float(total_viruksham_value),
            "scheme_items": scheme_items,
            "overall_remarks": summary_data.get("remarks") or "None",
        }

        # 6b. Fall back to AnchorParser key-value fields when employee aggregation yields zeros
        if not total_gold and summary_data.get("gold_sales"):
            total_gold = float(summary_data["gold_sales"])
        if not total_diamond and summary_data.get("diamond_sales"):
            total_diamond = float(summary_data["diamond_sales"])
        if not total_platinum and summary_data.get("platinum_sales"):
            total_platinum = float(summary_data["platinum_sales"])
        if not total_silver and summary_data.get("silver_sales"):
            total_silver = float(summary_data["silver_sales"])
        if not total_revenue and summary_data.get("total_revenue"):
            total_revenue = float(summary_data["total_revenue"])
        if not total_digigold and summary_data.get("digigold_enrollments"):
            total_digigold = int(summary_data["digigold_enrollments"])
        if not total_digisilver and summary_data.get("digisilver_enrollments"):
            total_digisilver = int(summary_data["digisilver_enrollments"])
        if not total_silver_mrp and summary_data.get("silver_mrp"):
            total_silver_mrp = float(summary_data["silver_mrp"])

        # 7. Run validation diagnostics
        diagnostics = ExtractionValidator.validate_and_build_diagnostics(
            summary_data=summary_data,
            employees_data=employees_data,
            scheme_data=scheme_summary,
            branch_totals={
                "gold": total_gold,
                "diamond": total_diamond,
                "platinum": total_platinum,
                "silver": total_silver,
                "silver_mrp": total_silver_mrp,
                "total_revenue": total_revenue,
                "digigold": total_digigold,
                "digisilver": total_digisilver,
            }
        )

        return {
            "summary": {
                "report_date": summary_data.get("report_date"),
                "branch_name": summary_data.get("branch_name"),
                "manager_name": summary_data.get("manager_name"),
                "sub_manager_name": summary_data.get("sub_manager_name"),
                # Branch-level aggregated values from employee rows
                "gold": float(total_gold),
                "diamond": float(total_diamond),
                "platinum": float(total_platinum),
                "silver": float(total_silver),
                "silver_mrp": float(total_silver_mrp),
                "total_revenue": float(total_revenue),
                "digigold": int(total_digigold),
                "digisilver": int(total_digisilver),
                # Metadata fields
                "employees_present": int(summary_data.get("employees_present") or 0),
                "employees_absent": int(summary_data.get("employees_absent") or 0),
                "customer_complaints": summary_data.get("customer_complaints") or "None",
                "operational_issues": summary_data.get("operational_issues") or "None",
                "remarks": summary_data.get("remarks") or "None",
            },
            "employees": employees_data,
            "scheme_summary": scheme_summary,
            "diagnostics": diagnostics,
        }
