import io
import re
import datetime
import openpyxl
from typing import Any, Optional

class SheetMatrix:
    """Wrapper around openpyxl Worksheet to transparently resolve merged cells and values."""
    def __init__(self, worksheet: openpyxl.worksheet.worksheet.Worksheet):
        self.worksheet = worksheet
        self.title = worksheet.title
        self.max_row = worksheet.max_row or 1
        self.max_column = worksheet.max_column or 1
        self._merged_map = {}
        self._build_merged_map()

    def _build_merged_map(self):
        """Map every coordinate inside a merged range to the top-left cell coordinate."""
        for rng in self.worksheet.merged_cells.ranges:
            top_left_val = self.worksheet.cell(row=rng.min_row, column=rng.min_col).value
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    self._merged_map[(r, c)] = top_left_val

    def get_cell_value(self, row: int, col: int) -> Any:
        """Get cell value, resolving merged cells if applicable."""
        if (row, col) in self._merged_map:
            return self._merged_map[(row, col)]
        return self.worksheet.cell(row=row, column=col).value


def clean_number(val: Any) -> float:
    """Cast cell value to float, handling currency symbols, words (INR, Rs), commas, and unicode."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    
    s = str(val).strip()
    # Strip currency words/symbols and commas
    s_clean = re.sub(r"(?i)\b(inr|rs|rupees|inr\.)\b|[₹$,/\-]", "", s).replace(",", "").strip()
    try:
        return float(s_clean)
    except ValueError:
        # Regex search for numeric float pattern
        match = re.search(r"[-+]?\d*\.?\d+", s.replace(",", ""))
        if match:
            try:
                return float(match.group(0))
            except ValueError:
                pass
        return 0.0


def clean_int(val: Any) -> int:
    """Cast cell value to int."""
    f_val = clean_number(val)
    return int(round(f_val))


def clean_string(val: Any, default: str = "None") -> str:
    """Clean string cell value."""
    if val is None:
        return default
    s = str(val).strip()
    if not s or s.lower() in ["none", "null", "n/a", "na", "-"]:
        return default
    return s


def clean_date(val: Any) -> Optional[str]:
    """Format date cell value to ISO string YYYY-MM-DD or return None if invalid."""
    if val is None:
        return None
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
        
    # Attempt to parse common date formats: YYYY-MM-DD, DD-MM-YYYY, YYYY/MM/DD, DD/MM/YYYY
    # Also strip any prefix like "Date:" or "As on:"
    s_clean = re.sub(r"(?i)\b(date|as on|on|dated|report date)\b\.?[\s:=]*", "", s).strip()
    
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%Y"):
        try:
            dt = datetime.datetime.strptime(s_clean, fmt)
            return dt.date().strftime("%Y-%m-%d")
        except ValueError:
            pass
            
    # Try regex match for date parts (e.g. 24-07-2026 or 2026-07-24)
    match = re.search(r"(\d{1,4})[\-\/](\d{1,2})[\-\/](\d{1,4})", s)
    if match:
        p1, p2, p3 = match.groups()
        for parts in [((p1, p2, p3), "%Y-%m-%d"), ((p1, p2, p3), "%d-%m-%Y"), ((p3, p2, p1), "%Y-%m-%d")]:
            try:
                y, m, d = parts[0]
                if len(y) == 2:
                    y = "20" + y
                if len(m) == 1:
                    m = "0" + m
                if len(d) == 1:
                    d = "0" + d
                dt = datetime.datetime.strptime(f"{y}-{m}-{d}", "%Y-%m-%d")
                return dt.date().strftime("%Y-%m-%d")
            except ValueError:
                pass

    # Try numeric excel serial number
    try:
        excel_date = float(s)
        dt = datetime.datetime(1899, 12, 30) + datetime.timedelta(days=excel_date)
        return dt.date().strftime("%Y-%m-%d")
    except ValueError:
        pass
        
    return None

